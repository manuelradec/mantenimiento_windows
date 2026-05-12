"""Tests for the T-04 history xlsx export.

Cubre:
- export_history_to_xlsx con DB vacía → produce xlsx válido con 5 hojas y headers.
- Con datos seeded → cuenta filas correcta y respeta filtros (date/hostname).
- Endpoint /reports/api/download/historico-xlsx responde xlsx + Content-Disposition.
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ============================================================
# Local app/client fixtures
# ============================================================


@pytest.fixture
def app(tmp_path):
    import config as cfg

    cfg.Config.LOG_DIR = str(tmp_path / "logs")
    cfg.Config.REPORT_DIR = str(tmp_path / "reports")
    os.makedirs(cfg.Config.LOG_DIR, exist_ok=True)
    os.makedirs(cfg.Config.REPORT_DIR, exist_ok=True)

    from app import create_app
    from core.persistence import init_db, SessionStore

    a = create_app()
    a.config["TESTING"] = True
    a.config["SESSION_ID"] = "test-session"
    a.config["HOSTNAME"] = "testhost"
    a.config["USERNAME"] = "testuser"
    init_db()
    SessionStore.create(
        session_id="test-session",
        hostname="testhost",
        username="testuser",
        is_admin=False,
    )
    return a


@pytest.fixture
def client(app):
    return app.test_client()


# ============================================================
# export_history_to_xlsx — service-level
# ============================================================


class TestExportHistoryXlsxEmpty:
    """Con DB casi vacía (solo la sesión 'test-session' del fixture)."""

    def test_returns_success_status(self, app):
        from services.maintenance_report import export_history_to_xlsx

        result = export_history_to_xlsx()
        assert result["status"] == "success"
        assert os.path.exists(result["path"])
        assert result["filename"].startswith("historico_mantenimientos_")
        assert result["filename"].endswith(".xlsx")

    def test_has_five_sheets_with_headers(self, app):
        from services.maintenance_report import export_history_to_xlsx
        from openpyxl import load_workbook

        result = export_history_to_xlsx()
        wb = load_workbook(result["path"], read_only=True)
        expected = [
            "Sessions",
            "Maintenance Steps",
            "Jobs",
            "Snapshots",
            "ScheduledRestarts",
        ]
        assert wb.sheetnames == expected

        # Cada hoja debe tener al menos una fila de headers.
        for name in expected:
            ws = wb[name]
            first_row = next(ws.iter_rows(values_only=True), None)
            assert first_row is not None, f"{name} sin header"
            assert len(first_row) >= 6, f"{name} header con pocas columnas"

    def test_session_fixture_appears_in_sessions_sheet(self, app):
        from services.maintenance_report import export_history_to_xlsx
        from openpyxl import load_workbook

        result = export_history_to_xlsx()
        wb = load_workbook(result["path"], read_only=True)
        ws = wb["Sessions"]
        rows = list(ws.iter_rows(values_only=True))
        # Header + al menos la sesión 'test-session' del fixture.
        assert len(rows) >= 2
        session_ids = [r[0] for r in rows[1:]]
        assert "test-session" in session_ids


# ============================================================
# export_history_to_xlsx — con seeded data
# ============================================================


class TestExportHistoryXlsxSeeded:
    def _seed(self):
        """Inserta filas de prueba en varias tablas."""
        from core.persistence import (
            SessionStore,
            AuditStore,
            JobStore,
            SnapshotStore,
            ScheduledRestartStore,
        )

        SessionStore.create(
            session_id="sess-A",
            hostname="HOST-A",
            username="userA",
            is_admin=False,
        )
        SessionStore.create(
            session_id="sess-B",
            hostname="HOST-B",
            username="userB",
            is_admin=True,
        )

        AuditStore.log(
            session_id="sess-A",
            module="maintenance",
            action="step_ccleaner",
            action_id="maintenance.ccleaner",
            status="completed",
            hostname="HOST-A",
            username="userA",
            duration_ms=12345,
            stdout_preview="OK",
            details={"freed_mb": 42},
        )
        AuditStore.log(
            session_id="sess-B",
            module="maintenance",
            action="step_defrag",
            action_id="maintenance.defrag",
            status="failed",
            hostname="HOST-B",
            username="userB",
            duration_ms=999,
        )
        # Ruido: entrada NO maintenance (no debe aparecer en Maintenance Steps).
        AuditStore.log(
            session_id="sess-A",
            module="network",
            action="flush_dns",
            action_id="network.flush_dns",
            status="completed",
            hostname="HOST-A",
        )

        JobStore.create(
            job_id="job-1",
            session_id="sess-A",
            action_id="cleanup.user_temp",
            action_name="Clean Temp",
            module="cleanup",
            risk_class="safe_mutation",
            hostname="HOST-A",
            username="userA",
        )

        SnapshotStore.save(
            job_id="job-1",
            session_id="sess-A",
            action_id="cleanup.user_temp",
            snapshot_type="before",
            captured_at="2099-01-01T10:00:00",
            data_json='{"disk_free_gb": 50}',
        )

        ScheduledRestartStore.record_create(
            scheduled_at="2099-06-01 22:00",
            recurrence="Once",
            grace_period=5,
            force=False,
            success=True,
            error="",
            session_id="sess-A",
            username="userA",
        )

    def test_seeded_data_appears_in_all_sheets(self, app):
        from services.maintenance_report import export_history_to_xlsx
        from openpyxl import load_workbook

        self._seed()
        result = export_history_to_xlsx()
        wb = load_workbook(result["path"], read_only=True)

        # Sessions: fixture's test-session + sess-A + sess-B = 3.
        sess_rows = list(wb["Sessions"].iter_rows(values_only=True))
        assert len(sess_rows) >= 4  # header + 3

        # Maintenance Steps: solo las 2 con module='maintenance' (no network).
        maint_rows = list(wb["Maintenance Steps"].iter_rows(values_only=True))
        actions = [r[3] for r in maint_rows[1:]]
        assert "step_ccleaner" in actions
        assert "step_defrag" in actions
        assert "flush_dns" not in actions  # ruido filtrado por module

        # Jobs: job-1.
        job_rows = list(wb["Jobs"].iter_rows(values_only=True))
        job_ids = [r[0] for r in job_rows[1:]]
        assert "job-1" in job_ids

        # Snapshots: 1.
        snap_rows = list(wb["Snapshots"].iter_rows(values_only=True))
        assert len(snap_rows) >= 2  # header + 1

        # ScheduledRestarts: 1.
        sr_rows = list(wb["ScheduledRestarts"].iter_rows(values_only=True))
        ops = [r[2] for r in sr_rows[1:]]
        assert "create" in ops

    def test_filter_by_hostname(self, app):
        from services.maintenance_report import export_history_to_xlsx
        from openpyxl import load_workbook

        self._seed()
        result = export_history_to_xlsx(hostname="HOST-A")
        wb = load_workbook(result["path"], read_only=True)

        sess_rows = list(wb["Sessions"].iter_rows(values_only=True))
        hosts = [r[3] for r in sess_rows[1:]]
        assert "HOST-A" in hosts
        assert "HOST-B" not in hosts

        maint_rows = list(wb["Maintenance Steps"].iter_rows(values_only=True))
        maint_hosts = [r[7] for r in maint_rows[1:]]
        assert "HOST-B" not in maint_hosts

    def test_filter_by_date_range_excludes_old(self, app):
        """date_from='2199-01-01' filtra todo a futuro → ningún match en
        las tablas con timestamp en presente."""
        from services.maintenance_report import export_history_to_xlsx
        from openpyxl import load_workbook

        self._seed()
        result = export_history_to_xlsx(date_from="2199-01-01")
        wb = load_workbook(result["path"], read_only=True)

        for sheet in ("Sessions", "Maintenance Steps", "Jobs"):
            rows = list(wb[sheet].iter_rows(values_only=True))
            # Solo el header — todo lo demás filtrado por date_from futuro.
            assert len(rows) == 1, f"{sheet} debería tener solo header"


# ============================================================
# Endpoint — flask client
# ============================================================


class TestHistoricoXlsxEndpoint:
    def test_endpoint_returns_xlsx_attachment(self, client):
        resp = client.get("/reports/api/download/historico-xlsx")
        assert resp.status_code == 200
        ct = resp.headers.get("Content-Type", "")
        assert (
            "spreadsheetml" in ct
            or "xlsx" in ct.lower()
            or ct.startswith("application/vnd.openxmlformats")
        )
        cd = resp.headers.get("Content-Disposition", "")
        assert "attachment" in cd.lower()
        assert "historico_mantenimientos_" in cd

    def test_endpoint_with_query_filters(self, client):
        resp = client.get(
            "/reports/api/download/historico-xlsx"
            "?hostname=testhost&date_from=2020-01-01"
        )
        assert resp.status_code == 200
