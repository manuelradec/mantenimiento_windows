"""
Maintenance Report Generation and Distribution.

Handles:
- HTML report generation
- Google Sheets integration (via gspread)
- Network share file copy
- RADEC Excel form FO-TI-19 generation (via openpyxl)
"""
import os
import re
import sys
import shutil
import logging
from datetime import datetime

from config import Config

logger = logging.getLogger('cleancpu.maintenance_report')

# Target paths
GOOGLE_SHEET_ID = '1i1v67mXuVA5Aqo2slYkrhLi_fDeUb95q'
NETWORK_SHARE_BASE = r'\\192.168.122.215\soporte CLJ\Mantenimiento Anual'
CREDENTIALS_PATH = os.path.join(Config.BASE_PATH, 'credentials', 'service_account.json')


def generate_html_report(system_info, steps, session_data):
    """Generate an HTML maintenance report."""
    hostname = system_info.get('hostname', 'UNKNOWN')
    serial = system_info.get('serial', 'UNKNOWN')
    date_str = datetime.now().strftime('%Y-%m-%d')
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    completed = sum(1 for s in steps if s.get('status') == 'completed')
    skipped = sum(1 for s in steps if s.get('status') == 'skipped')
    failed = sum(1 for s in steps if s.get('status') == 'failed')
    total_time = sum(s.get('elapsed', 0) for s in steps)
    status_text = 'COMPLETADO' if failed == 0 else 'PARCIAL'

    steps_html = ''
    for i, step in enumerate(steps, 1):
        status_color = {
            'completed': '#0AAE6B',
            'skipped': '#D6814A',
            'failed': '#E33B14',
            'cancelled': '#888',
        }.get(step.get('status', ''), '#666')
        steps_html += f'''
        <tr>
            <td style="padding:6px 10px;">{i}</td>
            <td style="padding:6px 10px;">{step.get('name', '')}</td>
            <td style="padding:6px 10px;color:{status_color};font-weight:bold;">{step.get('status', 'N/A').upper()}</td>
            <td style="padding:6px 10px;">{step.get('elapsed', 0):.1f}s</td>
            <td style="padding:6px 10px;font-size:12px;">{step.get('message', '')}</td>
        </tr>'''

    html = f'''<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Mantenimiento - {hostname} - {date_str}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 30px; color: #333; }}
        h1 {{ color: #274C9B; border-bottom: 3px solid #0AAE6B; padding-bottom: 8px; }}
        h2 {{ color: #0E7C5A; margin-top: 24px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 12px 0; }}
        th {{ background: #274C9B; color: #fff; padding: 8px 10px; text-align: left; }}
        td {{ border-bottom: 1px solid #ddd; }}
        .info-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 8px 24px; }}
        .info-grid dt {{ font-weight: bold; color: #555; }}
        .info-grid dd {{ margin: 0; }}
        .status {{ font-size: 18px; font-weight: bold; color: {'#0AAE6B' if failed == 0 else '#E33B14'}; }}
        .footer {{ margin-top: 30px; font-size: 12px; color: #999; border-top: 1px solid #ddd; padding-top: 10px; }}
    </style>
</head>
<body>
    <h1>Reporte de Mantenimiento Lógico — RADEC AUTOPARTES</h1>
    <p class="status">Estado: {status_text}</p>
    <p>Fecha: {timestamp}</p>

    <h2>Identificación del equipo</h2>
    <dl class="info-grid">
        <dt>Hostname:</dt><dd>{hostname}</dd>
        <dt>IP:</dt><dd>{system_info.get('ip_address', 'N/A')}</dd>
        <dt>No. Serie:</dt><dd>{serial}</dd>
        <dt>Marca:</dt><dd>{system_info.get('manufacturer', 'N/A')}</dd>
        <dt>Modelo:</dt><dd>{system_info.get('model', 'N/A')}</dd>
        <dt>Procesador:</dt><dd>{system_info.get('processor', 'N/A')}</dd>
        <dt>RAM:</dt><dd>{system_info.get('ram_gb', 'N/A')}</dd>
        <dt>Disco:</dt><dd>{system_info.get('hard_drive', 'N/A')}</dd>
        <dt>Sistema operativo:</dt><dd>{system_info.get('os_version', 'N/A')}</dd>
    </dl>

    <h2>Pasos del mantenimiento</h2>
    <table>
        <thead>
            <tr><th>#</th><th>Paso</th><th>Estado</th><th>Tiempo</th><th>Detalle</th></tr>
        </thead>
        <tbody>{steps_html}</tbody>
    </table>

    <h2>Resumen</h2>
    <ul>
        <li>Completados: {completed}/{len(steps)}</li>
        <li>Omitidos: {skipped}</li>
        <li>Fallidos: {failed}</li>
        <li>Tiempo total: {total_time:.0f} segundos</li>
    </ul>

    {_build_inventory_section_html(system_info)}

    <div class="footer">
        Generado por CleanCPU v{Config.APP_VERSION} — RADEC AUTOPARTES<br>
        {timestamp}
    </div>
</body>
</html>'''

    return html


def _build_inventory_section_html(system_info: dict) -> str:
    """
    Build an HTML inventory/system-snapshot section from system_info dict.
    Tolerant of missing keys — any absent field renders as 'N/A'.
    system_info may contain a nested 'inventory' key (from collect_inventory())
    or flat keys from the legacy _collect_system_info() call.
    """
    from html import escape as he

    inv = system_info.get('inventory', {})
    basic = inv.get('basic', {})
    hardware = inv.get('hardware', {})
    system = inv.get('system', {})
    network = inv.get('network', {})
    office = inv.get('office', {})

    def g(d: dict, key: str, fallback: str = 'N/A') -> str:
        """Get from dict, fall back to system_info flat keys, then fallback."""
        v = d.get(key) or system_info.get(key)
        return he(str(v).strip()) if v else fallback

    # ---- Basic ----
    rows_basic = [
        ('Fecha / Hora', f"{g(basic, 'date')} {g(basic, 'time')}"),
        ('Equipo (hostname)', g(basic, 'hostname', system_info.get('hostname', 'N/A'))),
        ('Usuario (login)', g(basic, 'username')),
        ('Nombre completo', g(basic, 'full_name')),
    ]

    # ---- Hardware ----
    rows_hw = [
        ('Fabricante', g(hardware, 'manufacturer', system_info.get('manufacturer', 'N/A'))),
        ('Modelo', g(hardware, 'model', system_info.get('model', 'N/A'))),
        ('No. de serie', g(hardware, 'serial', system_info.get('serial', 'N/A'))),
        ('UUID', g(hardware, 'uuid')),
        ('Dominio/Grupo', g(hardware, 'domain')),
        ('Tipo de union', g(hardware, 'join_type')),
    ]

    # ---- System ----
    ram_modules = system.get('ram_modules', [])
    ram_mod_str = ', '.join(
        f"{m.get('slot', '')} {m.get('capacity', '')} {m.get('type', '')} {m.get('speed', '')}"
        for m in ram_modules
    ) if ram_modules else ''

    disks = system.get('disks', [])
    disks_str = ' | '.join(
        f"{d.get('model', '')} {d.get('capacity', '')} ({d.get('media_type', '')})"
        for d in disks
    ) if disks else ''

    rows_sys = [
        ('Sistema operativo', g(system, 'os_name', system_info.get('os_version', 'N/A'))),
        ('Version', g(system, 'os_version')),
        ('Build', g(system, 'os_build')),
        ('Arquitectura', g(system, 'os_arch', system_info.get('architecture', 'N/A'))),
        ('Procesador', g(system, 'processor', system_info.get('processor', 'N/A'))),
        ('RAM total', g(system, 'ram_total')),
        ('Modulos RAM', he(ram_mod_str) if ram_mod_str else 'N/A'),
        ('Discos fisicos', he(disks_str) if disks_str else 'N/A'),
    ]

    # ---- Network ----
    rows_net = [
        ('Ethernet MAC', g(network, 'ethernet_mac')),
        ('Ethernet IPv4', g(network, 'ethernet_ip')),
        ('WiFi MAC', g(network, 'wifi_mac')),
        ('WiFi IPv4', g(network, 'wifi_ip')),
    ]

    # ---- Office ----
    rows_off = [
        ('Producto', g(office, 'product_name')),
        ('Version', g(office, 'version')),
        ('Plataforma', g(office, 'platform')),
        ('Canal', g(office, 'channel')),
        ('Release IDs', g(office, 'release_ids')),
    ]

    def _table(rows):
        return (
            '<table style="border-collapse:collapse;width:100%;margin:6px 0 14px;">'
            + ''.join(
                f'<tr><td style="padding:4px 10px;width:160px;font-weight:bold;'
                f'color:#555;">{k}</td>'
                f'<td style="padding:4px 10px;border-bottom:1px solid #eee;">{v}</td></tr>'
                for k, v in rows
            )
            + '</table>'
        )

    return f'''
    <h2 style="color:#0E7C5A;margin-top:28px;">Inventario del equipo</h2>

    <h3 style="margin:10px 0 4px;font-size:13px;color:#274C9B;">Identificación</h3>
    {_table(rows_basic)}

    <h3 style="margin:10px 0 4px;font-size:13px;color:#274C9B;">Hardware</h3>
    {_table(rows_hw)}

    <h3 style="margin:10px 0 4px;font-size:13px;color:#274C9B;">Sistema operativo</h3>
    {_table(rows_sys)}

    <h3 style="margin:10px 0 4px;font-size:13px;color:#274C9B;">Red</h3>
    {_table(rows_net)}

    <h3 style="margin:10px 0 4px;font-size:13px;color:#274C9B;">Microsoft Office</h3>
    {_table(rows_off)}
'''


def save_report_locally(html_content, system_info):
    """Save the HTML report to the local reports directory."""
    hostname = system_info.get('hostname', 'UNKNOWN')
    serial = system_info.get('serial', 'UNKNOWN')
    date_str = datetime.now().strftime('%Y-%m-%d')

    filename = f'Mantenimiento_{hostname}_{serial}_{date_str}.html'
    local_path = os.path.join(Config.REPORT_DIR, filename)

    try:
        os.makedirs(Config.REPORT_DIR, exist_ok=True)
        with open(local_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        logger.info(f"Report saved locally: {local_path}")
        return local_path
    except Exception as e:
        logger.error(f"Failed to save local report: {e}")
        return None


def save_to_network_share(html_content, system_info):
    """Save the report to the network share."""
    if sys.platform != 'win32':
        return {'status': 'skipped', 'reason': 'Only available on Windows'}

    hostname = system_info.get('hostname', 'UNKNOWN')
    serial = system_info.get('serial', 'UNKNOWN')
    date_str = datetime.now().strftime('%Y-%m-%d')
    filename = f'Mantenimiento_{hostname}_{serial}_{date_str}.html'

    target_dir = os.path.join(NETWORK_SHARE_BASE, date_str)
    target_path = os.path.join(target_dir, filename)

    try:
        os.makedirs(target_dir, exist_ok=True)
        with open(target_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        logger.info(f"Report saved to network share: {target_path}")
        return {'status': 'success', 'path': target_path}
    except (PermissionError, OSError) as e:
        logger.error(f"Failed to save to network share: {e}")
        return {
            'status': 'error',
            'error': f'No se pudo guardar en la carpeta de red: {e}. '
                     'El reporte se guardó localmente.',
        }


def update_google_sheets(system_info, steps):
    """Append a row to the Google Sheets maintenance log."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        logger.warning("gspread/google-auth not installed. Skipping Google Sheets update.")
        return {'status': 'skipped', 'reason': 'Bibliotecas de Google no instaladas.'}

    if not os.path.exists(CREDENTIALS_PATH):
        logger.warning(f"Google credentials not found at {CREDENTIALS_PATH}")
        return {
            'status': 'skipped',
            'reason': 'Credenciales de Google no configuradas. '
                      'Coloque el archivo service_account.json en la carpeta credentials/.',
        }

    try:
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive',
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=scopes)
        gc = gspread.authorize(creds)
        sheet = gc.open_by_key(GOOGLE_SHEET_ID).sheet1

        failed = sum(1 for s in steps if s.get('status') == 'failed')
        status_text = 'COMPLETADO' if failed == 0 else 'PARCIAL'

        details = '; '.join(
            f"{s.get('name', '')}: {s.get('status', '')}"
            for s in steps
        )

        row = [
            system_info.get('ip_address', 'N/A'),
            system_info.get('serial', 'N/A'),
            system_info.get('manufacturer', 'N/A'),
            system_info.get('model', 'N/A'),
            system_info.get('processor', 'N/A'),
            system_info.get('ram_gb', 'N/A'),
            system_info.get('hard_drive', 'N/A'),
            status_text,
            datetime.now().strftime('%d/%m/%Y'),
            details[:500],
        ]

        sheet.append_row(row, value_input_option='USER_ENTERED')
        logger.info("Google Sheets updated successfully")
        return {'status': 'success'}
    except Exception as e:
        logger.error(f"Google Sheets update failed: {e}")
        return {'status': 'error', 'error': str(e)}


def generate_radec_excel(system_info, steps):
    """
    Generate the RADEC FO-TI-19 Excel form.
    If a template exists, use it; otherwise create from scratch.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed. Skipping Excel form generation.")
        return {'status': 'skipped', 'reason': 'openpyxl no está instalado.'}

    hostname = system_info.get('hostname', 'UNKNOWN')
    date_str = datetime.now().strftime('%Y-%m-%d')
    date_display = datetime.now().strftime('%d/%m/%Y')

    # Try to load template
    template_path = os.path.join(Config.BASE_PATH, 'templates_data', 'FO-TI-19_template.xlsx')
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'FO-TI-19'
        _build_form_from_scratch(ws, system_info, steps, date_display)

    # Fill common fields regardless of template presence
    _fill_form_fields(ws, system_info, steps, date_display)

    # Save
    filename = f'FO-TI-19_{hostname}_{date_str}.xlsx'
    local_path = os.path.join(Config.REPORT_DIR, filename)
    os.makedirs(Config.REPORT_DIR, exist_ok=True)

    try:
        wb.save(local_path)
        logger.info(f"RADEC Excel form saved: {local_path}")
    except Exception as e:
        logger.error(f"Failed to save Excel form: {e}")
        return {'status': 'error', 'error': str(e)}

    # Also copy to network share
    network_result = {'status': 'skipped'}
    if sys.platform == 'win32':
        try:
            net_dir = os.path.join(NETWORK_SHARE_BASE, date_str)
            os.makedirs(net_dir, exist_ok=True)
            net_path = os.path.join(net_dir, filename)
            shutil.copy2(local_path, net_path)
            network_result = {'status': 'success', 'path': net_path}
            logger.info(f"Excel form copied to network: {net_path}")
        except (PermissionError, OSError) as e:
            network_result = {'status': 'error', 'error': str(e)}
            logger.warning(f"Failed to copy Excel to network: {e}")

    return {
        'status': 'success',
        'local_path': local_path,
        'network': network_result,
    }


def _build_form_from_scratch(ws, system_info, steps, date_display):
    """Build FO-TI-19 form structure when no template exists."""
    from openpyxl.styles import Font, Alignment, PatternFill

    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14, color='274C9B')
    fill_header = PatternFill(start_color='274C9B', end_color='274C9B', fill_type='solid')
    font_white = Font(bold=True, color='FFFFFF', size=10)

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'HOJA DE SERVICIO MANTENIMIENTO DE EQUIPO DE CÓMPUTO'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header info
    ws['A2'] = 'Código:'
    ws['B2'] = 'FO-TI-19'
    ws['D2'] = 'Versión: 06/07'
    ws['F2'] = f'Fecha de Emisión: {date_display}'

    # Equipment section
    row = 4
    ws[f'A{row}'] = 'DATOS DEL EQUIPO'
    ws[f'A{row}'].font = header_font

    labels = [
        ('Hostname:', system_info.get('hostname', '')),
        ('Procesador:', system_info.get('processor', '')),
        ('RAM:', system_info.get('ram_gb', '')),
        ('Disco Duro:', system_info.get('hard_drive', '')),
        ('Sistema Operativo:', system_info.get('os_version', '')),
        ('No. Serie:', system_info.get('serial', '')),
        ('Marca:', system_info.get('manufacturer', '')),
        ('Modelo:', system_info.get('model', '')),
        ('IP:', system_info.get('ip_address', '')),
    ]
    for i, (label, value) in enumerate(labels):
        r = row + 1 + i
        ws[f'A{r}'] = label
        ws[f'A{r}'].font = Font(bold=True, size=10)
        ws[f'B{r}'] = str(value)

    # Service type
    row = 15
    ws[f'A{row}'] = 'TIPO DE SERVICIO'
    ws[f'A{row}'].font = header_font
    ws[f'A{row + 1}'] = 'Preventivo'
    ws[f'B{row + 1}'] = 'X'
    ws[f'B{row + 1}'].font = Font(bold=True, color='0AAE6B')

    # Activities
    row = 18
    ws[f'A{row}'] = 'ACTIVIDADES REALIZADAS'
    ws[f'A{row}'].font = header_font

    # Table header
    row += 1
    for col, header in enumerate(['#', 'Actividad', 'Estado', 'Tiempo', 'Detalle'], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font_white
        cell.fill = fill_header

    # Steps
    for i, step in enumerate(steps, 1):
        r = row + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=step.get('name', ''))
        ws.cell(row=r, column=3, value=(step.get('status', '')).upper())
        ws.cell(row=r, column=4, value=f"{step.get('elapsed', 0):.1f}s")
        ws.cell(row=r, column=5, value=step.get('message', ''))

    # Observations
    obs_row = row + len(steps) + 2
    ws[f'A{obs_row}'] = 'OBSERVACIONES DEL TÉCNICO'
    ws[f'A{obs_row}'].font = header_font

    completed = sum(1 for s in steps if s.get('status') == 'completed')
    failed = sum(1 for s in steps if s.get('status') == 'failed')
    total_time = sum(s.get('elapsed', 0) for s in steps)

    ws[f'A{obs_row + 1}'] = (
        f'Mantenimiento lógico preventivo ejecutado por CleanCPU v{Config.APP_VERSION}. '
        f'{completed}/{len(steps)} pasos completados, {failed} fallidos. '
        f'Tiempo total: {total_time:.0f}s.'
    )

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 50


def _fill_form_fields(ws, system_info, steps, date_display):
    """Try to fill known cells in an existing template by searching for labels."""
    # Simple cell-value fill for known positions
    # This is a best-effort approach that works whether or not a template was loaded
    field_map = {
        'Fecha de Solicitud': date_display,
        'Fecha de Emisión': date_display,
    }

    for row in ws.iter_rows(max_row=min(50, ws.max_row or 50)):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip()
                if val in field_map:
                    # Fill adjacent cell
                    next_col = cell.column + 1
                    ws.cell(row=cell.row, column=next_col, value=field_map[val])


def generate_fo_ti_19_html(system_info, steps, session_data,
                           sucursal='', technician_name='',
                           maint_type='preventivo', model_override='',
                           tech_address='', tech_phone='', tech_email='',
                           operator_name='', op_address='', op_phone='', op_email='',
                           accessories_override='', drive_overrides=None,
                           office_license=None):
    """
    Generate FO-TI-19 Hoja de Servicio Mantenimiento de Equipo de Cómputo.
    Matches the official RADEC format with all fields.
    """
    from html import escape as he

    date_display = datetime.now().strftime('%d/%m/%Y')

    # Service type checkmarks (derived from maint_type)
    _mt = (maint_type or 'preventivo').lower()
    svc_preventivo = 'checked' if _mt == 'preventivo' else ''
    svc_correctivo = 'checked' if _mt == 'correctivo' else ''
    svc_revision = 'checked' if _mt == 'revision' else ''

    # Determine accessories
    accessories = (
        accessories_override.strip()
        if accessories_override and accessories_override.strip()
        else 'MOUSE, TECLADO Y NO-BREAK'
    )

    # Unidades (drives) checkmarks
    has_cdrom = system_info.get('has_cdrom', False)
    has_dvdrom = system_info.get('has_dvdrom', False)
    has_usb = system_info.get('has_usb', True)
    has_micro_sd = system_info.get('has_micro_sd', False)

    # Otros text
    otros_text = ''
    otros_parts = []
    if has_usb:
        otros_parts.append('USB')
    if has_micro_sd:
        otros_parts.append('MICRO SD')
    if otros_parts:
        otros_text = ', '.join(otros_parts)

    # Build drive-type override note for HD row
    _drive_note = ''
    if drive_overrides and isinstance(drive_overrides, dict):
        _types = [str(v).strip().upper() for v in drive_overrides.values()
                  if v and str(v).strip()]
        if _types:
            _drive_note = ' — Tipos confirmados: ' + ', '.join(_types)

    # Activities text (based on maint_type)
    _activities_map = {
        'preventivo': (
            'SE REALIZO LA LIMPIEZA INTERNA COMO EXTERNA, ASI COMO EL '
            'MANTENIMIENTO LOGICO, SE ENTREGA EQUIPO FUNCIONANDO CORRECTAMENTE.'
        ),
        'correctivo': (
            'SE REALIZO EL DIAGNOSTICO Y CORRECCION DE FALLA REPORTADA. '
            'SE ENTREGA EQUIPO FUNCIONANDO CORRECTAMENTE.'
        ),
        'revision': (
            'SE REALIZO REVISION GENERAL DEL EQUIPO. '
            'SE ENTREGA INFORME DE ESTADO.'
        ),
    }
    activities = _activities_map.get(_mt, _activities_map['preventivo'])

    # Physical observations
    observations_physical = 'EQUIPO FUNCIONAL'

    # Comments (based on maint_type)
    _comments_map = {
        'preventivo': 'MANTENIMIENTO ANUAL PREVENTIVO',
        'correctivo': 'MANTENIMIENTO CORRECTIVO',
        'revision': 'REVISION DE EQUIPO',
    }
    comments = _comments_map.get(_mt, _comments_map['preventivo'])

    # Technician observations
    tech_observations = ''
    if steps:
        step_details = []
        for s in steps:
            if s.get('status') == 'failed':
                step_details.append(f"{s.get('name', '')}: FALLIDO - {s.get('message', '')}")
            elif s.get('status') == 'skipped':
                step_details.append(f"{s.get('name', '')}: OMITIDO - {s.get('message', '')}")
        if step_details:
            tech_observations = '; '.join(step_details)
        else:
            tech_observations = 'TODOS LOS PASOS COMPLETADOS EXITOSAMENTE'

    # Technician contact sub-lines for signature block
    _tech_contact = ''
    if tech_address:
        _tech_contact += (
            f'<br><span style="font-size:9px;font-weight:normal;">{he(tech_address)}</span>'
        )
    if tech_phone or tech_email:
        _contact_line = ' | '.join(filter(None, [tech_phone, tech_email]))
        _tech_contact += (
            f'<br><span style="font-size:9px;font-weight:normal;">{he(_contact_line)}</span>'
        )

    # Upgrade opportunities / Mejoras section
    upgrades = system_info.get('upgrade_opportunities', {})
    cpu_obs = _check_cpu_obsolescence(system_info.get('processor', ''))
    mejoras_html = _build_mejoras_section(upgrades, cpu_obsolescence=cpu_obs)

    # Office license block (from session inspection result)
    office_html = _build_office_license_section(office_license)

    html = f'''<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>FO-TI-19 - {he(system_info.get('hostname', 'EQUIPO'))}</title>
    <style>
        @page {{ size: letter; margin: 15mm; }}
        @media print {{ body {{ margin: 0; }} .no-print {{ display: none; }} }}
        body {{ font-family: Arial, sans-serif; font-size: 11px; color: #000; margin: 20px; }}
        .form-container {{ max-width: 800px; margin: 0 auto; border: 2px solid #000; }}
        .header {{ display: flex; align-items: center; border-bottom: 2px solid #000; }}
        .header-logo {{ width: 150px; padding: 8px 12px; border-right: 1px solid #000;
                        font-weight: bold; font-size: 18px; color: #C00; text-align: center; }}
        .header-logo .sub {{ font-size: 8px; color: #666; letter-spacing: 2px; }}
        .header-title {{ flex: 1; text-align: center; padding: 8px; }}
        .header-title h2 {{ margin: 0; font-size: 14px; }}
        .header-title .doc-type {{ font-size: 9px; color: #666; }}
        .header-code {{ width: 130px; border-left: 1px solid #000; padding: 6px 10px;
                        font-size: 10px; }}
        .header-code div {{ margin: 2px 0; }}
        .row {{ display: flex; border-bottom: 1px solid #000; }}
        .cell {{ padding: 6px 10px; border-right: 1px solid #000; }}
        .cell:last-child {{ border-right: none; }}
        .cell-label {{ font-weight: bold; background: #f5f5f5; min-width: 140px; font-size: 10px; }}
        .cell-value {{ flex: 1; font-size: 11px; }}
        .section-title {{ background: #e8e8e8; font-weight: bold; padding: 6px 10px;
                          border-bottom: 1px solid #000; font-size: 11px; }}
        .checkbox {{ display: inline-block; width: 14px; height: 14px; border: 1.5px solid #000;
                     text-align: center; line-height: 14px; font-size: 10px; margin-right: 4px;
                     vertical-align: middle; }}
        .checkbox.checked {{ background: #000; color: #fff; }}
        .radio {{ display: inline-block; width: 14px; height: 14px; border: 1.5px solid #000;
                  border-radius: 50%; text-align: center; line-height: 14px; font-size: 8px;
                  margin-right: 4px; vertical-align: middle; }}
        .radio.checked {{ background: #000; color: #fff; }}
        .signature-area {{ display: flex; border-bottom: 1px solid #000; min-height: 70px; }}
        .signature-box {{ flex: 1; border-right: 1px solid #000; padding: 10px; text-align: center;
                          display: flex; flex-direction: column; justify-content: flex-end; }}
        .signature-box:last-child {{ border-right: none; }}
        .signature-line {{ border-top: 1px solid #000; margin-top: auto; padding-top: 4px;
                           font-size: 10px; font-weight: bold; }}
        .mejoras-section {{ margin-top: 0; border-top: 2px solid #000; }}
        .mejoras-section h3 {{ background: #274C9B; color: #fff; margin: 0; padding: 8px 10px;
                               font-size: 12px; }}
        .mejora-item {{ padding: 6px 10px; border-bottom: 1px solid #ddd; font-size: 11px; }}
        .mejora-item .tag {{ display: inline-block; background: #0AAE6B; color: #fff;
                             padding: 1px 6px; border-radius: 3px; font-size: 9px;
                             font-weight: bold; margin-right: 6px; }}
        .mejora-item .tag.storage {{ background: #D6814A; }}
        .mejora-item .tag.nvme {{ background: #274C9B; }}
        .mejora-item .tag.obsolete {{ background: #C00; }}
        .mejora-item .tag.manual {{ background: #888; }}
        .footer-version {{ text-align: right; font-size: 9px; color: #999; padding: 4px 10px; }}
    </style>
</head>
<body>
<div class="form-container">
    <!-- Header -->
    <div class="header">
        <div class="header-logo">
            RADEC<br><span class="sub">AUTOPARTES</span>
        </div>
        <div class="header-title">
            <h2>Hoja de Servicio Mantenimiento de Equipo de C&oacute;mputo</h2>
            <div class="doc-type">Tipo de Documento<br>Formato</div>
        </div>
        <div class="header-code">
            <div><strong>C&oacute;digo</strong></div>
            <div>FO-TI-19</div>
            <div><strong>Fecha de Emisi&oacute;n</strong></div>
            <div>10/01/2025</div>
        </div>
    </div>

    <!-- Fecha de Solicitud -->
    <div class="row">
        <div class="cell cell-label" style="flex:1;">
            <strong>Fecha de Solicitud:</strong>
        </div>
        <div class="cell cell-value" style="flex:1;text-align:center;font-weight:bold;">
            {he(date_display)}
        </div>
    </div>

    <!-- Sucursal -->
    <div class="row">
        <div class="cell cell-label">Sucursal:</div>
        <div class="cell cell-value">{he(sucursal)}</div>
    </div>

    <!-- Nombre del Solicitante -->
    <div class="row">
        <div class="cell cell-label">Nombre del Solicitante:</div>
        <div class="cell cell-value" style="text-align:center;font-weight:bold;">
            {he((operator_name or system_info.get('user_fullname', system_info.get('username', ''))).upper()) or '&nbsp;'}
        </div>
    </div>

    <!-- Direccion -->
    <div class="row">
        <div class="cell cell-label">Direcci&oacute;n:</div>
        <div class="cell cell-value">{he(op_address) or '&nbsp;'}</div>
    </div>

    <!-- Telefono / Correo -->
    <div class="row">
        <div class="cell cell-label" style="width:100px;">Tel&eacute;fono:</div>
        <div class="cell cell-value" style="flex:1;">{he(op_phone) or '&nbsp;'}</div>
        <div class="cell cell-label" style="width:130px;">Correo Electr&oacute;nico:</div>
        <div class="cell cell-value" style="flex:1;">
            {he(op_email or system_info.get('user_email', '')) or '&nbsp;'}
        </div>
    </div>

    <!-- Descripcion del Equipo -->
    <div class="row">
        <div class="cell cell-label">Descripci&oacute;n del Equipo:</div>
        <div class="cell cell-value">
            {he(model_override or system_info.get('equipment_description',
                system_info.get('model', 'N/A')))}
        </div>
    </div>

    <!-- Datos del Monitor -->
    <div class="row">
        <div class="cell cell-label">Datos del Monitor:</div>
        <div class="cell cell-value">{he(system_info.get('monitor_info', 'N/A'))}</div>
    </div>

    <!-- Procesador, Velocidad -->
    <div class="row">
        <div class="cell cell-label">Procesador, Velocidad:</div>
        <div class="cell cell-value">{he(system_info.get('processor', 'N/A'))}</div>
    </div>

    <!-- Capacidad de RAM / Discos de Sistema -->
    <div class="row">
        <div class="cell cell-label" style="width:140px;">Capacidad de RAM:</div>
        <div class="cell cell-value" style="flex:1;">
            {he(system_info.get('ram_detail', system_info.get('ram_gb', 'N/A')))}
        </div>
        <div class="cell cell-label" style="width:130px;">Discos de Sistema:</div>
        <div class="cell cell-value" style="width:80px;">
            <span class="radio checked">&#8226;</span> S&iacute;
            &nbsp;&nbsp;
            <span class="radio">&#8226;</span> No
        </div>
    </div>

    <!-- Capacidad de HD / Sistema Operativo -->
    <div class="row">
        <div class="cell cell-label" style="width:140px;">Capacidad de HD:</div>
        <div class="cell cell-value" style="flex:1;">
            {he(system_info.get('hard_drive', 'N/A'))}{he(_drive_note)}
        </div>
        <div class="cell cell-label" style="width:130px;">Sistema Operativo:</div>
        <div class="cell cell-value" style="flex:1;">
            {he(system_info.get('os_version_full', system_info.get('os_version', 'N/A')))}
        </div>
    </div>

    <!-- Unidades -->
    <div class="row">
        <div class="cell cell-label" style="width:140px;">Unidades:</div>
        <div class="cell cell-value" style="flex:1;">
            <div style="margin-bottom:4px;">
                <span class="checkbox {'checked' if has_cdrom else ''}">{('X' if has_cdrom else '&nbsp;')}</span> CD-ROM
                &nbsp;&nbsp;&nbsp;
                <span class="checkbox">{'X' if False else '&nbsp;'}</span> Quemador CD
                &nbsp;&nbsp;&nbsp;
                <span class="checkbox">{'X' if False else '&nbsp;'}</span> Flopy
                &nbsp;&nbsp;&nbsp;
                <span class="checkbox {'checked' if otros_text else ''}">{('X' if otros_text else '&nbsp;')}</span> Otros
                &nbsp; <span style="font-size:10px;">{he(otros_text)}</span>
            </div>
            <div>
                <span class="checkbox {'checked' if has_dvdrom else ''}">{('X' if has_dvdrom else '&nbsp;')}</span> DVD-ROM
                &nbsp;&nbsp;&nbsp;
                <span class="checkbox">{'X' if False else '&nbsp;'}</span> Quemador DVD
            </div>
        </div>
    </div>

    <!-- Accesorios -->
    <div class="row">
        <div class="cell cell-label">Accesorios:</div>
        <div class="cell cell-value" style="text-align:center;">{he(accessories)}</div>
    </div>

    <!-- Tipo de Servicio Solicitado -->
    <div class="row">
        <div class="cell cell-label" style="width:140px;">Tipo de Servicio<br>Solicitado:</div>
        <div class="cell cell-value" style="flex:1;display:flex;justify-content:space-around;align-items:center;">
            <span><span class="radio {'checked' if svc_revision else ''}">&#8226;</span> Revisi&oacute;n</span>
            <span><span class="radio {'checked' if svc_preventivo else ''}">&#8226;</span> Preventivo</span>
            <span><span class="radio {'checked' if svc_correctivo else ''}">&#8226;</span> Correctivo</span>
        </div>
    </div>

    <!-- Comentarios sobre la falla -->
    <div class="row">
        <div class="cell cell-label">Comentarios sobre la<br>falla:</div>
        <div class="cell cell-value" style="text-align:center;min-height:40px;display:flex;align-items:center;justify-content:center;">
            {he(comments)}
        </div>
    </div>

    <!-- Observaciones Estado Fisico -->
    <div class="row">
        <div class="cell cell-label">Observaciones Estado<br>F&iacute;sico:</div>
        <div class="cell cell-value" style="text-align:center;min-height:40px;display:flex;align-items:center;justify-content:center;">
            {he(observations_physical)}
        </div>
    </div>

    <!-- Signatures -->
    <div class="signature-area">
        <div class="signature-box">
            <div style="min-height:40px;"></div>
            <div class="signature-line">
                {he((operator_name or system_info.get('user_fullname', '')).upper()) or '&nbsp;'}<br>
                Nombre y Firma del Solicitante
            </div>
        </div>
        <div class="signature-box">
            <div style="min-height:40px;"></div>
            <div class="signature-line">
                {he(technician_name.upper()) if technician_name else '&nbsp;'}<br>
                Nombre y Firma de quien Recibe{_tech_contact}
            </div>
        </div>
    </div>

    <!-- Actividades Realizadas -->
    <div class="row" style="min-height:60px;">
        <div class="cell cell-label">Actividades Realizadas:</div>
        <div class="cell cell-value" style="text-align:center;display:flex;align-items:center;justify-content:center;">
            {he(activities)}
        </div>
    </div>

    <!-- Observaciones del Tecnico -->
    <div class="row" style="min-height:40px;">
        <div class="cell cell-label">Observaciones del<br>T&eacute;cnico:</div>
        <div class="cell cell-value" style="text-align:center;display:flex;align-items:center;justify-content:center;">
            {he(tech_observations)}
        </div>
    </div>

    <!-- MEJORAS Section -->
    {mejoras_html}

    <!-- LICENCIA OFFICE Section -->
    {office_html}

</div>

<div class="footer-version">
    Versi&oacute;n 07 &mdash; P&aacute;gina 1 de 1 &mdash;
    Generado por CleanCPU v{Config.APP_VERSION}
</div>

<div class="no-print" style="text-align:center;margin-top:20px;">
    <button onclick="window.print()" style="padding:10px 30px;font-size:14px;cursor:pointer;">
        Imprimir / Guardar PDF
    </button>
</div>
</body>
</html>'''
    return html


def generate_fo_ti_20_html(entries, sucursal=''):
    """
    Generate FO-TI-20 Bitacora de mantenimiento de equipo de computo.
    entries: list of dicts with keys: fecha, usuario, equipo, reporte_final
    Matches the official RADEC summary table format.
    """
    from html import escape as he

    date_display = datetime.now().strftime('%d/%m/%Y')

    rows_html = ''
    for i, entry in enumerate(entries, 1):
        rows_html += f'''
        <tr>
            <td style="text-align:center;">{i}</td>
            <td style="text-align:center;">{he(str(entry.get('fecha', '')))}</td>
            <td style="text-align:center;">{he(str(entry.get('usuario', '')).upper())}</td>
            <td style="text-align:center;">{he(str(entry.get('equipo', '')).upper())}</td>
            <td style="text-align:center;">{he(str(entry.get('reporte_final', 'MANTENIMIENTO PREVENTIVO')).upper())}</td>
            <td style="text-align:center;min-width:80px;">&nbsp;</td>
        </tr>'''

    # Fill remaining empty rows to reach at least 20
    for j in range(len(entries) + 1, 21):
        rows_html += f'''
        <tr>
            <td style="text-align:center;">{j}</td>
            <td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td>
        </tr>'''

    html = f'''<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>FO-TI-20 Bit&aacute;cora de Mantenimiento - {he(sucursal)}</title>
    <style>
        @page {{ size: letter landscape; margin: 15mm; }}
        @media print {{ body {{ margin: 0; }} .no-print {{ display: none; }} }}
        body {{ font-family: Arial, sans-serif; font-size: 11px; color: #000; margin: 20px; }}
        .form-container {{ max-width: 1000px; margin: 0 auto; border: 2px solid #000; }}
        .header {{ display: flex; align-items: center; border-bottom: 2px solid #000; }}
        .header-logo {{ width: 150px; padding: 8px 12px; border-right: 1px solid #000;
                        font-weight: bold; font-size: 18px; color: #C00; text-align: center; }}
        .header-logo .sub {{ font-size: 8px; color: #666; letter-spacing: 2px; }}
        .header-title {{ flex: 1; text-align: center; padding: 8px; }}
        .header-title h2 {{ margin: 0; font-size: 14px; }}
        .header-title .doc-type {{ font-size: 9px; color: #666; }}
        .header-code {{ width: 130px; border-left: 1px solid #000; padding: 6px 10px;
                        font-size: 10px; }}
        .header-code div {{ margin: 2px 0; }}
        .sucursal-row {{ display: flex; border-bottom: 2px solid #000; }}
        .sucursal-label {{ padding: 8px 12px; font-weight: bold; font-size: 12px; flex: 1; }}
        .sucursal-value {{ padding: 8px 12px; font-weight: bold; font-size: 12px; flex: 1;
                           text-align: right; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th {{ background: #e8e8e8; border: 1px solid #000; padding: 8px 6px; font-size: 11px;
              font-weight: bold; }}
        td {{ border: 1px solid #000; padding: 6px; font-size: 11px; }}
        .footer-version {{ text-align: right; font-size: 9px; color: #999; padding: 4px 10px; }}
    </style>
</head>
<body>
<div class="form-container">
    <!-- Header -->
    <div class="header">
        <div class="header-logo">
            RADEC<br><span class="sub">AUTOPARTES</span>
        </div>
        <div class="header-title">
            <h2>Bit&aacute;cora de mantenimiento de equipo de computo</h2>
            <div class="doc-type">Tipo de Documento<br>Formato</div>
        </div>
        <div class="header-code">
            <div><strong>C&oacute;digo</strong></div>
            <div>FO-TI-20</div>
            <div><strong>Fecha de Emisi&oacute;n</strong></div>
            <div>10/01/2025</div>
        </div>
    </div>

    <!-- Sucursal -->
    <div class="sucursal-row">
        <div class="sucursal-label">Sucursal</div>
        <div class="sucursal-value">{he(sucursal.upper())}</div>
    </div>

    <!-- Table -->
    <table>
        <thead>
            <tr>
                <th style="width:30px;">#</th>
                <th style="width:100px;">Fecha</th>
                <th style="width:200px;">Usuario</th>
                <th style="width:200px;">Equipo</th>
                <th style="width:200px;">Reporte Final</th>
                <th style="width:100px;">Firma</th>
            </tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
    </table>
</div>

<div class="footer-version">
    Generado por CleanCPU v{Config.APP_VERSION} &mdash; {he(date_display)}
</div>

<div class="no-print" style="text-align:center;margin-top:20px;">
    <button onclick="window.print()" style="padding:10px 30px;font-size:14px;cursor:pointer;">
        Imprimir / Guardar PDF
    </button>
</div>
</body>
</html>'''
    return html


def _check_cpu_obsolescence(processor_str):
    """
    Option A obsolescence rule:
    A machine is a mandatory renewal candidate if its processor is BELOW
    Intel Core i5 5th generation.

    Threshold matrix:
      - Gen 6+, any Core family  → NOT obsolete
      - Gen 5 + Core i5/i7/i9   → NOT obsolete  (at threshold or above)
      - Gen 5 + Core i3          → obsolete       (below threshold)
      - Gen 1-4, any Core family → obsolete

    AMD, Xeon, Pentium, Celeron, Atom, and unrecognized families are never
    silently skipped; all are reported as "Revisión manual requerida".

    Returns dict:
      generation (int or None), family (str or None, e.g. 'I5'),
      obsolete  (True / False / None),
      label     (str — display name),
      reason    (str — explanation for the report)
    """
    if not processor_str:
        return {
            'generation': None, 'family': None, 'obsolete': None,
            'label': 'Procesador no identificado',
            'reason': 'Revisi\u00f3n manual requerida.',
        }

    p = processor_str.upper()

    # Intel Core iX with explicit model number: iN-DDDD[suffix]
    m = re.search(r'CORE\s+(I[3579])-(\d{4,5})', p)
    if m:
        family = m.group(1)                              # 'I3', 'I5', 'I7', 'I9'
        fdisp = family[0].lower() + family[1:]          # 'i3', 'i5', 'i7', 'i9'
        model_num = m.group(2)

        if len(model_num) == 5:
            # Gen 10+: first two digits (10xxx → 10, 11xxx → 11, …)
            gen = int(model_num[:2])
        else:
            # Gen 1-9: first digit (4xxx → 4, 5xxx → 5, …)
            gen = int(model_num[0]) or 1

        label = f'Intel Core {fdisp} \u2014 Generaci\u00f3n {gen}'

        if gen >= 6:
            obsolete = False
            reason = (
                f'Generaci\u00f3n {gen}, dentro del umbral de soporte '
                f'(m\u00ednimo requerido: Core i5, 5\u00aa generaci\u00f3n).'
            )
        elif gen == 5:
            if family == 'I3':
                # i3 5th gen is below the threshold (i5 5th gen)
                obsolete = True
                reason = (
                    'Core i3 de 5\u00aa generaci\u00f3n no alcanza el umbral m\u00ednimo. '
                    'Se requiere Core i5 de 5\u00aa generaci\u00f3n o superior.'
                )
            else:
                # i5/i7/i9 on 5th gen — at or above threshold
                obsolete = False
                reason = (
                    f'Core {fdisp} de 5\u00aa generaci\u00f3n, en umbral m\u00ednimo o superior '
                    f'(m\u00ednimo requerido: Core i5, 5\u00aa generaci\u00f3n).'
                )
        else:
            # gen <= 4: all Core families are below threshold
            obsolete = True
            reason = (
                f'Generaci\u00f3n {gen}, por debajo del umbral m\u00ednimo '
                f'(se requiere Core i5 de 5\u00aa generaci\u00f3n o superior).'
            )

        return {
            'generation': gen, 'family': family,
            'obsolete': obsolete, 'label': label, 'reason': reason,
        }

    # Intel Core 1st-generation — old WMI naming without hyphen:
    # "Intel(R) Core(TM) i5 CPU 760 @ 2.80GHz"  (desktop, 3-digit)
    # "Intel(R) Core(TM) i5 CPU M 520 @ 2.40GHz" (mobile with letter suffix)
    # Identified by: family letter + whitespace + "CPU" keyword + 3-digit number.
    # All CPUs matching this style are definitively 1st generation → obsolete.
    m1g = re.search(r'\b(I[3579])\s+CPU\s+(?:[A-Z]\s+)?\d{3}\b', p)
    if m1g:
        family = m1g.group(1)                        # 'I3', 'I5', 'I7'
        fdisp = family[0].lower() + family[1:]      # 'i3', 'i5', 'i7'
        return {
            'generation': 1, 'family': family, 'obsolete': True,
            'label': f'Intel Core {fdisp} \u2014 Generaci\u00f3n 1',
            'reason': (
                'Generaci\u00f3n 1 (modelo de 3 d\u00edgitos, denominaci\u00f3n '
                'antigua sin gui\u00f3n), por debajo del umbral m\u00ednimo '
                '(se requiere Core i5 de 5\u00aa generaci\u00f3n o superior).'
            ),
        }

    # Intel Core iX present but no model number — generation indeterminate
    if re.search(r'CORE\s+I[3579]\b', p):
        fm = re.search(r'CORE\s+(I[3579])\b', p)
        fdisp = (fm.group(1)[0].lower() + fm.group(1)[1:]) if fm else 'iX'
        return {
            'generation': None, 'family': fdisp, 'obsolete': None,
            'label': f'Intel Core {fdisp} (sin n\u00famero de modelo)',
            'reason': 'Generaci\u00f3n no determinada. Revisi\u00f3n manual requerida.',
        }

    # Known non-Core-iX families — rule does not apply; manual review required
    _FAMILY_MAP = [
        ('XEON', 'Intel Xeon'),
        ('CELERON', 'Intel Celeron'),
        ('PENTIUM', 'Intel Pentium'),
        ('ATOM', 'Intel Atom'),
        ('RYZEN', 'AMD Ryzen'),
        ('ATHLON', 'AMD Athlon'),
        ('EPYC', 'AMD EPYC'),
        ('THREADRIPPER', 'AMD Threadripper'),
        ('AMD', 'AMD'),
        ('APPLE', 'Apple Silicon'),
    ]
    for keyword, brand in _FAMILY_MAP:
        if keyword in p:
            return {
                'generation': None, 'family': brand, 'obsolete': None,
                'label': brand,
                'reason': 'Familia no clasificable por esta regla. Revisi\u00f3n manual requerida.',
            }

    # Completely unrecognized string
    return {
        'generation': None, 'family': None, 'obsolete': None,
        'label': 'Procesador no reconocido',
        'reason': 'Familia de procesador no reconocida. Revisi\u00f3n manual requerida.',
    }


def _build_office_license_section(office_license):
    """
    Build the Office License HTML section for FO-TI-19.

    Reads only safe, pre-masked fields from the session inspection result:
      - product_name, license_status, partial_key (last-5 only), inspected_at
    The full product key is never stored in the session and is never shown here.

    Returns '' (empty string, section omitted) only when office_license is None.
    When it is a dict but fields are missing, the section still renders with
    whatever partial data is available and clearly flags missing fields.
    """
    from html import escape as he

    # Always render the section so the technician can see whether inspection ran.
    # Fallback block when no inspection was performed this session.
    if not office_license or not isinstance(office_license, dict):
        return '''
<div class="section-title">LICENCIA OFFICE</div>
<div class="row">
    <div class="cell cell-value" style="color:#888;font-style:italic;padding:8px 10px;">
        Licencia Office: No inspeccionada en esta sesi&oacute;n.
        Ejecute &ldquo;Inspeccionar licencia&rdquo; en la secci&oacute;n Office
        y regenere el reporte.
    </div>
</div>'''

    # Extract safe fields defensively — all defaults to '' if absent or None
    parsed = office_license.get('parsed', {}) or {}
    product_name = he(str(parsed.get('product_name', '') or ''))
    license_status = he(str(parsed.get('license_status', '') or ''))
    partial_key = he(str(parsed.get('partial_key', '') or ''))
    inspected_at_raw = str(office_license.get('inspected_at', '') or '')
    status = str(office_license.get('status', '') or '')
    message = he(str(office_license.get('message', '') or ''))

    # Format inspected_at for display (ISO → readable, best-effort)
    if inspected_at_raw:
        try:
            from datetime import datetime as _dt
            _ts = _dt.fromisoformat(inspected_at_raw)
            inspected_at = _ts.strftime('%d/%m/%Y %H:%M')
        except (ValueError, TypeError):
            inspected_at = he(inspected_at_raw[:16])
    else:
        inspected_at = ''

    # Status badge color
    _status_colors = {
        'success': '#006600',
        'requires_admin': '#CC6600',
        'ospp_not_found': '#888888',
        'office_not_found': '#888888',
        'error': '#CC0000',
    }
    status_color = _status_colors.get(status, '#333333')

    # Build display rows — only show rows that have data
    rows = []

    if product_name:
        rows.append(
            f'<div class="row">'
            f'<div class="cell cell-label" style="min-width:160px;">Producto:</div>'
            f'<div class="cell cell-value">{product_name}</div>'
            f'</div>'
        )

    if license_status:
        rows.append(
            f'<div class="row">'
            f'<div class="cell cell-label" style="min-width:160px;">Estado de licencia:</div>'
            f'<div class="cell cell-value" style="color:{status_color};font-weight:bold;">'
            f'{license_status}</div>'
            f'</div>'
        )

    if partial_key:
        rows.append(
            f'<div class="row">'
            f'<div class="cell cell-label" style="min-width:160px;">'
            f'&Uacute;ltimos 5 caracteres:</div>'
            f'<div class="cell cell-value" style="font-family:monospace;">'
            f'XXXXX-XXXXX-XXXXX-XXXXX-{partial_key}</div>'
            f'</div>'
        )

    if inspected_at:
        rows.append(
            f'<div class="row">'
            f'<div class="cell cell-label" style="min-width:160px;">Inspeccionado:</div>'
            f'<div class="cell cell-value" style="color:#555;font-size:10px;">'
            f'{inspected_at}</div>'
            f'</div>'
        )

    # When ospp ran but produced no parsed fields, show the summary message
    if not rows and message:
        rows.append(
            f'<div class="row">'
            f'<div class="cell cell-value" style="color:{status_color};padding:8px 10px;">'
            f'{message}</div>'
            f'</div>'
        )

    if not rows:
        rows.append(
            '<div class="row">'
            '<div class="cell cell-value" style="color:#888;font-style:italic;padding:8px 10px;">'
            'Sin datos de licencia disponibles.</div>'
            '</div>'
        )

    return (
        '<div class="section-title">LICENCIA OFFICE</div>\n'
        + '\n'.join(rows)
    )


def _build_mejoras_section(upgrades, cpu_obsolescence=None):
    """Build the Mejoras (Improvements) HTML section for FO-TI-19."""
    from html import escape as he

    # Skip only when there is neither hardware upgrade data nor a CPU result.
    # cpu_obsolescence is always shown (including manual-review results).
    _has_cpu = bool(cpu_obsolescence and
                    (cpu_obsolescence.get('label') or cpu_obsolescence.get('reason')))
    if not upgrades and not _has_cpu:
        return ''

    recommendations = upgrades.get('recommendations', [])
    ram_info = upgrades.get('ram', {})
    storage_info = upgrades.get('storage', {})
    expansion_info = upgrades.get('expansion', {})

    items_html = ''

    # RAM details
    if ram_info.get('empty_slots', 0) > 0:
        items_html += f'''
        <div class="mejora-item">
            <span class="tag">RAM</span>
            <strong>Slots disponibles:</strong> {ram_info.get('empty_slots', 0)} de
            {ram_info.get('total_slots', 0)} total.
            Capacidad actual: {ram_info.get('current_capacity_gb', 0):.0f} GB /
            M&aacute;ximo soportado: {ram_info.get('max_capacity_gb', 0)} GB.
        </div>'''
        # Show installed modules
        for mod in ram_info.get('modules', []):
            cap = mod.get('CapacityGB', 0) or 0
            mfr = mod.get('Manufacturer', 'N/A') or 'N/A'
            speed = mod.get('ConfiguredClockSpeed', 0) or mod.get('Speed', 0) or 0
            bank = mod.get('BankLabel', '') or mod.get('DeviceLocator', '') or ''
            items_html += f'''
            <div class="mejora-item" style="padding-left:30px;font-size:10px;">
                Slot {he(str(bank))}: {cap:.0f} GB - {he(str(mfr).strip())} @ {speed} MHz
            </div>'''
    elif ram_info.get('total_slots', 0) > 0:
        items_html += f'''
        <div class="mejora-item">
            <span class="tag">RAM</span>
            Todos los slots ocupados ({ram_info.get('occupied_slots', 0)} de
            {ram_info.get('total_slots', 0)}).
            Capacidad actual: {ram_info.get('current_capacity_gb', 0):.0f} GB.
            {'Se pueden reemplazar m&oacute;dulos por unos de mayor capacidad.' if ram_info.get('current_capacity_gb', 0) < ram_info.get('max_capacity_gb', 0) else 'Capacidad m&aacute;xima alcanzada.'}
        </div>'''

    # Storage details
    if storage_info.get('has_hdd', False):
        items_html += '''
        <div class="mejora-item">
            <span class="tag storage">HDD</span>
            <strong>Disco mec&aacute;nico detectado.</strong>
            Se recomienda actualizar a SSD/NVMe para mejorar rendimiento.
        </div>'''

    for disk in storage_info.get('disks', []):
        name = disk.get('FriendlyName', 'N/A') or 'N/A'
        media = disk.get('MediaType', 'N/A') or 'N/A'
        bus = disk.get('BusType', '') or ''
        size = disk.get('SizeGB', 0) or 0
        health = disk.get('HealthStatus', 'N/A') or 'N/A'
        items_html += f'''
        <div class="mejora-item" style="padding-left:30px;font-size:10px;">
            {he(str(name))} - {size:.0f} GB - Tipo: {he(str(media))}
            {f'({he(str(bus))})' if bus else ''} - Salud: {he(str(health))}
        </div>'''

    # NVMe/M.2 expansion
    if expansion_info.get('m2_slots_available', 0) > 0:
        items_html += f'''
        <div class="mejora-item">
            <span class="tag nvme">M.2</span>
            <strong>{expansion_info.get('m2_slots_available', 0)} slot(s) M.2 disponible(s)</strong>
            de {expansion_info.get('m2_slots_total', 0)} total.
            Se puede instalar almacenamiento NVMe adicional.
        </div>'''
    elif expansion_info.get('m2_slots_total', 0) == 0 and not storage_info.get('has_nvme', False):
        items_html += '''
        <div class="mejora-item">
            <span class="tag nvme">M.2</span>
            No se detectaron slots M.2/NVMe disponibles.
        </div>'''

    # General recommendations
    for rec in recommendations:
        if rec not in items_html:  # Avoid duplicates
            items_html += f'''
            <div class="mejora-item">
                &#9679; {he(rec)}
            </div>'''

    # CPU obsolescence block — rendered for all results, including manual review
    if _has_cpu:
        obs = cpu_obsolescence
        reason_text = obs.get('reason', '')
        if obs.get('obsolete') is True:
            tag_class = 'obsolete'
            msg = (f'<strong>Candidato a renovaci\u00f3n:</strong> '
                   f'{he(obs.get("label", ""))} \u2014 {he(reason_text)}')
        elif obs.get('obsolete') is False:
            tag_class = ''
            msg = f'{he(obs.get("label", ""))} \u2014 {he(reason_text)}'
        else:
            tag_class = 'manual'
            lbl = obs.get('label', '')
            msg = he(lbl) + (f' \u2014 {he(reason_text)}' if reason_text else '')

        items_html += f'''
        <div class="mejora-item">
            <span class="tag cpu {tag_class}">CPU</span>
            {msg}
        </div>'''

    if not items_html:
        items_html = '''
        <div class="mejora-item">
            No se detectaron oportunidades de mejora de hardware.
            El equipo est&aacute; en su capacidad &oacute;ptima.
        </div>'''

    return f'''
    <div class="mejoras-section">
        <h3>MEJORAS DE HARDWARE DETECTADAS</h3>
        {items_html}
    </div>'''


def generate_full_report(session_data):
    """
    Generate all reports after maintenance completion.
    Returns a summary of all report generation attempts.
    """
    from routes.maintenance import _collect_system_info
    system_info = _collect_system_info()
    steps = session_data.get('steps', [])

    # Attach full inventory snapshot to system_info for report sections.
    # Collected independently — a failure here does not abort report generation.
    try:
        from services.system_inventory import collect_inventory
        system_info['inventory'] = collect_inventory()
    except Exception as e:
        logger.warning(f"Inventory collection failed during report generation: {e}")
        system_info['inventory'] = {}

    results = {}

    # 1. HTML report
    html = generate_html_report(system_info, steps, session_data)
    local_path = save_report_locally(html, system_info)
    results['html_local'] = (
        {'status': 'success', 'path': local_path} if local_path
        else {'status': 'error'}
    )

    # 2. Network share
    results['network_share'] = save_to_network_share(html, system_info)

    # 3. Google Sheets
    results['google_sheets'] = update_google_sheets(system_info, steps)

    # 4. RADEC Excel form
    results['excel_form'] = generate_radec_excel(system_info, steps)

    # 5. FO-TI-19 HTML (Hoja de Servicio)
    try:
        fo_ti_19_html = generate_fo_ti_19_html(
            system_info, steps, session_data
        )
        hostname = system_info.get('hostname', 'UNKNOWN')
        date_str = datetime.now().strftime('%Y-%m-%d')
        filename = f'FO-TI-19_{hostname}_{date_str}.html'
        filepath = os.path.join(Config.REPORT_DIR, filename)
        os.makedirs(Config.REPORT_DIR, exist_ok=True)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(fo_ti_19_html)
        results['fo_ti_19'] = {'status': 'success', 'path': filepath}
        logger.info(f"FO-TI-19 HTML saved: {filepath}")
    except Exception as e:
        logger.error(f"FO-TI-19 generation failed: {e}")
        results['fo_ti_19'] = {'status': 'error', 'error': str(e)}

    return results
